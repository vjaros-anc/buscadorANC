# -*- coding: utf-8 -*-
"""
Extrae los productos de las definiciones de mercado de firm4.xlsx y los deja
escritos como columnas del propio Excel, para poder corregirlos a mano.

Antes, los productos se calculaban al vuelo dentro de generar_pagina.py cada
vez que se armaba el index. Ahora el circuito es:

    extraer_productos.py  ->  columnas nuevas en firm4.xlsx  ->  (edicion
    manual en Excel)  ->  nomenclador_mercados.py / generar_pagina.py  ->  index

Columnas que agrega (al final de la hoja, si no existen):

  productos            Productos detectados en el texto de mercado relevante,
                       separados por " | ". Son las etiquetas del catalogo
                       PRODUCTOS de nomenclador_mercados.py.
  productos_sector     Sector(es) del nomenclador del expediente, separados por
                       " | ". Agrupa los productos en el desplegable del index
                       y alimenta el filtro por sector.
  productos_sugeridos  Frases del mercado relevante que NINGUN patron del
                       catalogo cubrio. Es material de revision: si alguna
                       vale como producto, se la escribe a mano en `productos`
                       (y, si conviene que se detecte sola en el futuro, se
                       agrega al dict PRODUCTOS del nomenclador).

REGLA DE ESCRITURA: solo se completan las celdas VACIAS. Todo lo que ya tenga
contenido se respeta, corrida tras corrida. Por eso se puede editar el Excel a
mano sin miedo a perder los cambios.

Si un expediente NO tiene que mostrar productos, no dejes la celda vacia (se
volveria a completar en la proxima corrida): escribi un guion "-". El extractor
lo respeta y el buscador lo ignora (ver MARCAS_VACIO en el nomenclador).

Texto de origen: se usan la V1 ("Mercados relevantes") y la V2
("mercado_relev_V2") en union -- lo que aparece en cualquiera de las dos entra.

Uso:
    python extraer_productos.py                     # completa las celdas vacias
    python extraer_productos.py --dry-run           # no escribe, solo informa
    python extraer_productos.py --refrescar-sugeridos   # recalcula esa columna
    python extraer_productos.py --excel otro.xlsx
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

import nomenclador_mercados as nm

AQUI = Path(__file__).parent
ARCHIVO = AQUI / "firm4.xlsx"
RESPALDOS = AQUI / "respaldos"

# nombres de columna en firm4.xlsx (origen)
C_CARATULA = "Carátula"
C_CARPETA = "Carpeta"
C_MERC_V1 = "Mercados relevantes"
C_MERC_V2 = "mercado_relev_V2"

# columnas que genera este script (destino)
C_PRODUCTOS = "productos"
C_PROD_SECTOR = "productos_sector"
C_PROD_SUGERIDOS = "productos_sugeridos"

SEP = " | "
SIN_SECTOR = "Otros / sin clasificar"

# frases sueltas que no aportan nada como producto sugerido
RUIDO_SUGERIDOS = re.compile(
    r"^(y|o|u|e|de|del|la|el|los|las|un|una|otros?|otras?|varios?|s/d|n/a|"
    r"idem|sin datos|no aplica|no corresponde)$"
)


def _p(*args) -> None:
    """print tolerante con consolas que no soportan acentos (cp850/cp437)."""
    txt = " ".join(str(a) for a in args)
    try:
        print(txt)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "ascii"
        print(txt.encode(enc, "replace").decode(enc))


# --------------------------------------------------------------------------- #
# Catalogo de productos (viene del nomenclador; unico lugar donde se edita)
# --------------------------------------------------------------------------- #
def catalogo() -> list[tuple[str, re.Pattern]]:
    """[(etiqueta, regex compilado), ...] a partir de nm.PRODUCTOS."""
    out: list[tuple[str, re.Pattern]] = []
    for etiqueta, info in nm.PRODUCTOS.items():
        patron = "|".join(info["patrones"])
        try:
            out.append((etiqueta, re.compile(patron)))
        except re.error as e:
            _p(f"  ! patron invalido en '{etiqueta}': {patron}  ({e})")
    return out


# --------------------------------------------------------------------------- #
# Extraccion por expediente
# --------------------------------------------------------------------------- #
def segmentos_union(merc_v1: str, merc_v2: str) -> list[str]:
    """Segmentos de la V1 y la V2 juntos, sin repetir (compara normalizado)."""
    vistos: set[str] = set()
    out: list[str] = []
    for texto in (merc_v1, merc_v2):
        for seg in nm.segmentar_mercados(texto):
            k = nm.norm(seg)
            if k and k not in vistos:
                vistos.add(k)
                out.append(seg)
    return out


def extraer_productos(segmentos: list[str], cat: list[tuple[str, re.Pattern]]) -> tuple[list[str], list[str]]:
    """Devuelve (productos detectados, frases sin cubrir).

    El texto contra el que se matchea es el mismo que usaba el index hasta
    ahora: los segmentos unidos y normalizados. La cobertura, en cambio, se
    evalua segmento por segmento, para poder senalar cual quedo sin producto.
    """
    if not segmentos:
        return [], []

    texto_norm = nm.norm(SEP.join(segmentos))
    productos = sorted({et for et, rx in cat if rx.search(texto_norm)}, key=nm.norm)

    sugeridos: list[str] = []
    for seg in segmentos:
        seg_norm = nm.norm(seg)
        if not seg_norm or RUIDO_SUGERIDOS.match(seg_norm):
            continue
        if any(rx.search(seg_norm) for _, rx in cat):
            continue
        sugeridos.append(seg.replace("|", "/").strip())

    return productos, sugeridos


def extraer_sectores(segmentos: list[str], caratula: str) -> list[str]:
    """Sectores del nomenclador para el expediente (mercado + caratula)."""
    base = nm.norm(SEP.join(segmentos) + " " + caratula)
    return nm.clasificar_sectores(base) or [SIN_SECTOR]


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def _mapa_columnas(ws) -> dict[str, int]:
    return {
        str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    }


def _asegurar_columna(ws, cols: dict[str, int], nombre: str) -> tuple[int, bool]:
    """Indice de la columna `nombre`; la crea al final si no existe."""
    if nombre in cols:
        return cols[nombre], False
    idx = ws.max_column + 1
    celda = ws.cell(row=1, column=idx, value=nombre)
    # copia el estilo del encabezado de al lado para que no desentone
    ref = ws.cell(row=1, column=max(1, idx - 1))
    celda.font = copy(ref.font)
    celda.fill = copy(ref.fill)
    celda.border = copy(ref.border)
    celda.alignment = copy(ref.alignment)
    ws.column_dimensions[celda.column_letter].width = 42
    cols[nombre] = idx
    return idx, True


def _vacia(valor) -> bool:
    return valor is None or not str(valor).strip() or str(valor).strip().lower() == "nan"


def _respaldar(ruta: Path) -> Path:
    RESPALDOS.mkdir(exist_ok=True)
    destino = RESPALDOS / f"{ruta.stem}_{datetime.now():%Y%m%d-%H%M%S}{ruta.suffix}"
    shutil.copy2(ruta, destino)
    return destino


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--excel", default=str(ARCHIVO), help="ruta del xlsx (default: firm4.xlsx de esta carpeta)")
    ap.add_argument("--dry-run", action="store_true", help="no escribe nada, solo informa que haria")
    ap.add_argument("--refrescar-sugeridos", action="store_true",
                    help="recalcula `productos_sugeridos` en TODAS las filas (pisa lo que haya)")
    ap.add_argument("--sin-respaldo", action="store_true", help="no deja copia en ./respaldos")
    args = ap.parse_args()

    ruta = Path(args.excel)
    if not ruta.is_absolute():
        ruta = AQUI / ruta
    if not ruta.exists():
        _p(f"No existe el archivo: {ruta}")
        sys.exit(1)

    cat = catalogo()
    _p(f"Catalogo: {len(cat)} productos (dict PRODUCTOS de nomenclador_mercados.py)")

    wb = openpyxl.load_workbook(ruta)
    ws = wb.worksheets[0]
    cols = _mapa_columnas(ws)

    faltan = [c for c in (C_CARATULA, C_MERC_V1, C_MERC_V2) if c not in cols]
    if faltan:
        _p("Faltan columnas de origen en el Excel:", ", ".join(faltan))
        _p("Columnas encontradas:", ", ".join(cols))
        sys.exit(1)

    c_prod, nuevo_prod = _asegurar_columna(ws, cols, C_PRODUCTOS)
    c_sec, nuevo_sec = _asegurar_columna(ws, cols, C_PROD_SECTOR)
    c_sug, nuevo_sug = _asegurar_columna(ws, cols, C_PROD_SUGERIDOS)
    for nombre, es_nueva in ((C_PRODUCTOS, nuevo_prod), (C_PROD_SECTOR, nuevo_sec),
                             (C_PROD_SUGERIDOS, nuevo_sug)):
        _p(f"  columna '{nombre}': {'creada' if es_nueva else 'ya existia'}")

    n_prod_esc = n_sec_esc = n_sug_esc = 0
    n_prod_resp = n_sec_resp = n_sug_resp = 0
    n_con_texto = n_sin_producto = 0
    universo: set[str] = set()
    sugeridos_todos: dict[str, int] = {}

    for fila in range(2, ws.max_row + 1):
        caratula = nm.clean(ws.cell(row=fila, column=cols[C_CARATULA]).value)
        carpeta = nm.clean(ws.cell(row=fila, column=cols[C_CARPETA]).value) if C_CARPETA in cols else ""
        merc_v1 = nm.clean(ws.cell(row=fila, column=cols[C_MERC_V1]).value)
        merc_v2 = nm.clean(ws.cell(row=fila, column=cols[C_MERC_V2]).value)
        if not (caratula or carpeta or merc_v1 or merc_v2):
            continue

        segmentos = segmentos_union(merc_v1, merc_v2)
        productos, sugeridos = extraer_productos(segmentos, cat)
        sectores = extraer_sectores(segmentos, caratula)

        if segmentos:
            n_con_texto += 1
            if not productos:
                n_sin_producto += 1
        universo.update(productos)
        for s in sugeridos:
            sugeridos_todos[s] = sugeridos_todos.get(s, 0) + 1

        # productos: solo si esta vacia
        cel = ws.cell(row=fila, column=c_prod)
        if _vacia(cel.value):
            if productos:
                if not args.dry_run:
                    cel.value = SEP.join(productos)
                n_prod_esc += 1
        else:
            n_prod_resp += 1

        # sector: solo si esta vacia
        cel = ws.cell(row=fila, column=c_sec)
        if _vacia(cel.value):
            if not args.dry_run:
                cel.value = SEP.join(sectores)
            n_sec_esc += 1
        else:
            n_sec_resp += 1

        # sugeridos: solo si esta vacia, salvo --refrescar-sugeridos
        cel = ws.cell(row=fila, column=c_sug)
        if args.refrescar_sugeridos:
            if not args.dry_run:
                cel.value = SEP.join(sugeridos) if sugeridos else None
            if sugeridos:
                n_sug_esc += 1
        elif _vacia(cel.value):
            if sugeridos:
                if not args.dry_run:
                    cel.value = SEP.join(sugeridos)
                n_sug_esc += 1
        else:
            n_sug_resp += 1

    # el autofiltro tiene que abarcar las columnas nuevas y las filas agregadas
    # desde la ultima corrida
    if ws.auto_filter.ref and not args.dry_run:
        inicio = ws.auto_filter.ref.split(":")[0]
        ultima = ws.cell(row=1, column=ws.max_column).column_letter
        ws.auto_filter.ref = f"{inicio}:{ultima}{ws.max_row}"

    if not args.dry_run:
        if not args.sin_respaldo:
            _p(f"Respaldo: {_respaldar(ruta).relative_to(AQUI)}")
        wb.save(ruta)

    modo = "DRY-RUN (no se escribio nada)" if args.dry_run else f"OK -> {ruta.name}"
    _p("")
    _p(modo)
    _p(f"  Expedientes con mercado relevante cargado: {n_con_texto}"
       f"  (sin ningun producto detectado: {n_sin_producto})")
    _p(f"  productos           : {n_prod_esc} celdas completadas | {n_prod_resp} respetadas (ya tenian contenido)")
    _p(f"  productos_sector    : {n_sec_esc} celdas completadas | {n_sec_resp} respetadas")
    _p(f"  productos_sugeridos : {n_sug_esc} celdas completadas | {n_sug_resp} respetadas"
       + ("  [refrescadas]" if args.refrescar_sugeridos else ""))
    _p(f"  Productos distintos en uso: {len(universo)} de {len(cat)} del catalogo")

    if sugeridos_todos:
        top = sorted(sugeridos_todos.items(), key=lambda kv: (-kv[1], nm.norm(kv[0])))
        _p(f"  Frases sin cubrir por el catalogo: {len(sugeridos_todos)} distintas. Las mas repetidas:")
        for frase, n in top[:15]:
            _p(f"    {n:3d}  {frase[:90]}")


if __name__ == "__main__":
    main()
