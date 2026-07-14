#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor de dictámenes CNDC (resolución + dictamen) -> planilla 'firmadas'.

Diseño: cada campo se extrae con un extractor independiente que devuelve None
si NO encuentra el dato de forma fiable. Un None se escribe como celda en blanco;
nunca se inventa ni se rellena con datos dudosos. Un PDF que falla no corta el lote.

Uso:
    python extraer_firmadas.py                 # procesa ./pdfs -> res_firmadas_nuevo.xlsx
    python extraer_firmadas.py <carpeta>       # carpeta con PDFs
    python extraer_firmadas.py <carpeta> <salida.xlsx>

Dependencias:
    pip install pymupdf openpyxl
"""
import sys, os, re, glob, unicodedata, datetime, json
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Configuración de columnas de salida (orden final de la planilla nueva)
#   - se AGREGA 'Grupo/Empresa'
#   - se ELIMINA 'DIAS'  (queda solo 'meses')
# ----------------------------------------------------------------------------
COLUMNS = [
    "tipo", "Carpeta", "Excluible", "Carátula", "Grupo/Empresa",
    "Empresas involucradas",
    "Fecha_Ingreso", "Fecha_firma", "Decisión",
    "Número de Resolución", "Número de Dictamen",
    "Mercados relevantes", "Relaciones económicas", "meses",
]
COL = {name: i + 1 for i, name in enumerate(COLUMNS)}  # 1-indexed

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

# ----------------------------------------------------------------------------
# Utilidades de texto
# ----------------------------------------------------------------------------
def leer_pdf(path):
    # sort=True reordena por posición: recupera la capa de encabezado GDE
    # (Número RESOL, "CIUDAD DE BUENOS AIRES" + fecha) en orden de lectura.
    # Además se reconstruyen las tablas por geometría (find_tables) para poder
    # leer la fila "Objeto" de la tabla de empresas involucradas.
    partes, tablas = [], []
    with fitz.open(path) as doc:
        for pg in doc:
            partes.append(pg.get_text(sort=True) or "")
            try:
                for tb in pg.find_tables().tables:
                    tablas.append(tb.extract())
            except Exception:
                pass
    raw = "\n".join(partes).replace("\xa0", " ")
    return raw, tablas

def flat(txt):
    """Colapsa saltos de línea/espacios para anclar sobre prosa."""
    return re.sub(r"\s+", " ", txt).strip()

def segmentar(raw):
    """Divide en (resolución, dictamen). Si no hay marca, ambos = documento completo."""
    idx = raw.find("Dictamen firma conjunta")
    if idx == -1:
        idx = raw.find("Dictamen firma")
    if idx > 0:
        return raw[:idx], raw[idx:]
    return raw, raw

def parse_fecha_es(dia, mes_txt, anio):
    m = MESES_ES.get(strip_acentos(mes_txt).lower())
    if not m:
        return None
    try:
        return datetime.date(int(anio), m, int(dia))
    except ValueError:
        return None

def strip_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")

def limpiar(s):
    # no se recorta el punto final para preservar "S.A." / "S.A.I.C."
    return re.sub(r"\s+", " ", s).strip(" \u201c\u201d\"'") if s else s

RE_FECHA = r"(\d{1,2}) de ([A-Za-zÁÉÍÓÚáéíóúÜü]+) de (\d{4})"

# ----------------------------------------------------------------------------
# Extractores  (cada uno devuelve str/date  ó  None)
# ----------------------------------------------------------------------------
def ex_tipo(ctx):
    return "CONC"

def ex_carpeta(ctx):
    m = re.search(r"CONC\.?\s*N?[.°º]?\s*(\d{3,4})", ctx["raw"])
    return f"CONC. {m.group(1)}" if m else None

def ex_caratula(ctx):
    m = re.search(r"caratulad[oa]s?\s*[\"«\u201c]([^\"»\u201d]{5,300})[\"»\u201d]",
                  ctx["dict_flat"], re.I)
    if not m:
        return None
    s = m.group(1)
    s = re.sub(r"^\s*CONC\.?\s*\d+\s*[-–]\s*", "", s)          # prefijo "CONC. 1844 -"
    s = re.sub(r"\s*[\(\[]\s*CONC\.?\s*\d+\s*[\)\]]\s*$", "", s)  # sufijo "(CONC 1859)"
    return limpiar(s)

def ex_grupo(ctx):
    car = ctx.get("caratula")
    if car:
        m = re.search(r"^(.*?)\s*S/\s*(?:NOTIFIC|SOLICIT)", car, re.I)
        if m and m.group(1).strip():
            return limpiar(m.group(1))
    # fallback: "adquisición por parte de [la firma] X del/de ..."
    m = re.search(r"adquisici[oó]n\s+(?:de control\s+)?por parte de\s+(?:la firma\s+)?"
                  r"([A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúñ0-9&.\-\s]+?)\s+"
                  r"(?:del|de la|de las|de los)\b", ctx["dict_flat"])
    return limpiar(m.group(1)) if m else None

def ex_fecha_ingreso(ctx):
    d = ctx["dict_flat"]
    for pat in (
        r"notificad[ao]s?\s+(?:con fecha\s+|el\s+)?" + RE_FECHA,
        r"[Cc]on fecha\s+" + RE_FECHA + r"[^.]{0,90}?(?:recibió la notificación|notific)",
        r"(?:recibió la notificación|notificaron)[^.]{0,90}?" + RE_FECHA,
    ):
        m = re.search(pat, d)
        if m:
            f = parse_fecha_es(m.group(1), m.group(2), m.group(3))
            if f:
                return f
    return None

def ex_fecha_firma(ctx):
    m = re.search(r"CIUDAD DE BUENOS AIRES\s+(?:[A-Za-zÁÉÍÓÚáéíóú]+\s+)?" + RE_FECHA,
                  ctx["reso_flat"])
    if m:
        return parse_fecha_es(m.group(1), m.group(2), m.group(3))
    return None

def ex_decision(ctx):
    # Se detecta sentido + inciso y se emite SIEMPRE la forma canónica de la hoja:
    #   "<Sentido> Art.14 <letra>) Ley 27.442"
    d = ctx["dict_flat"]
    sentido, letra = None, None
    m = re.search(r"(Autoriza|Rechaza|Subordina|Deniega|Condiciona)\s*Art\.?\s*14"
                  r"[^.\n]*?\(?\s*([abc])\s*\)", d, re.I)
    if m:
        sentido, letra = m.group(1).capitalize(), m.group(2).lower()
    else:
        inc = re.search(r"art[íi]culo\s*14\s*inc(?:iso)?\.?\s*\(?([abc])\)?", d, re.I)
        if inc:
            letra = inc.group(1).lower()
            sentido = {"a": "Autoriza", "b": "Subordina", "c": "Rechaza"}[letra]
    if sentido and letra:
        return f"{sentido} Art.14 {letra}) Ley 27.442"
    return None

def ex_num_resolucion(ctx):
    m = re.search(r"RESOL-\d{4}-\d+-APN-SC#MEC", ctx["raw"])
    return m.group(0) if m else None

def ex_num_dictamen(ctx):
    m = re.search(r"IF-\s*\d{4}-\s*\d+-APN-CNDC#MEC", ctx["dict_raw"]) \
        or re.search(r"IF-\s*\d{4}-\s*\d+-APN-CNDC#MEC", ctx["raw"])
    return re.sub(r"\s+", "", m.group(0)) if m else None

def _celda_numerica(s):
    s = (s or "").strip()
    return bool(s) and re.fullmatch(r"[\d\.,%\s\-\$]+", s) is not None

def _label_objeto(s):
    # admite "Objeto", "Empresa Objeto" y "Empresas Objeto" (singular/plural)
    return bool(re.match(r"^\s*(empresas?\s+)?objeto\b", s.strip(), re.I)) and "(" not in s

def _label_comprador(s):
    return bool(re.search(r"adquir|comprad|lado adquirente", s, re.I))

def ex_mercados(ctx):
    """Mercado relevante = descripción de la actividad del 'Objeto' en la
    tabla de empresas involucradas. El estado de sección se ARRASTRA entre
    tablas (el rótulo 'Objeto' y sus filas pueden quedar en tablas/páginas
    distintas); se corta en tablas de participaciones/IHH."""
    descs = []
    recolectando = False
    for tabla in ctx["tablas"]:
        for row in tabla:
            if not row:
                continue
            cells = [(c or "").strip() for c in row]
            first = cells[0]
            rest = cells[1:]
            rest_txt = " ".join(x for x in rest if x).strip()
            es_label = bool(first) and not rest_txt      # solo la 1ra celda con texto
            if es_label:
                recolectando = _label_objeto(first)      # True solo en sección Objeto
                continue
            if _es_fila_participacion(cells):            # corta al llegar a IHH/participaciones
                recolectando = False
                continue
            if recolectando:
                desc = " ".join(x for x in rest if x and not _celda_numerica(x)).strip()
                if desc and re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]{3,}", desc):
                    descs.append(re.sub(r"\s+", " ", desc).strip())
    if not descs:
        return None
    return "; ".join(descs)[:400].strip(" ;")

def ex_relaciones(ctx):
    d = ctx["dict_flat"]
    has_h = re.search(r"relaci[oó]n(?:es)? horizontal(?:es)?|"
                      r"presenta relaciones horizontales|"
                      r"fortalecimiento (?:del grupo|horizontal)|"
                      r"relaci[oó]n horizontal", d, re.I) is not None
    has_v = re.search(r"efectos? vertical(?:es)?|relaci[oó]n(?:es)? vertical(?:es)?|"
                      r"reforzamiento (?:de\s+)?(?:un |las |los )?(?:efectos?|relaci[oó]n(?:es)?)?\s*vertical",
                      d, re.I) is not None
    has_c = re.search(r"conglomerado", d, re.I) is not None
    partes = []
    if has_h: partes.append("h")
    if has_v: partes.append("v")
    if has_c: partes.append("c")
    if not partes:
        return None
    if partes == ["h"]:            return "Horizontal"
    if partes == ["v"]:            return "Vertical"
    if partes == ["c"]:            return "Conglomerado"
    if set(partes) == {"h", "v"}:  return "Horizontal + vertical"
    # combinaciones con conglomerado
    nombre = {"h": "Horizontal", "v": "Vertical", "c": "Conglomerado"}
    return " + ".join(nombre[p] for p in partes)

# ----------------------------------------------------------------------------
# Empresas involucradas (tabla de "empresas afectadas / involucradas")
#   Recorre las tablas en orden de lectura y devuelve la lista de empresas
#   preservando el orden del dictamen, cada una con su rol (comprador/objeto).
#   El rol surge de:
#     - etiquetas de sección: "Grupo comprador", "Objeto", "Adquirente",
#       "Empresas del Grupo Adquirente", "Activos Adquiridos", etc.
#     - marcadores entre paréntesis en la misma celda: "(objeto)", "(Grupo comprador)".
#     - "Grupo XXX" sin palabra clave -> se asume lado comprador (el objeto casi
#       siempre viene rotulado explícitamente).
#   Se descartan las tablas de participaciones de mercado / IHH (celdas numéricas,
#   años, "Participación", etc.) para no contaminar la lista.
#   Salida: JSON compacto  [{"rol": "...", "nombre": "..."}, ...]  ó None.
# ----------------------------------------------------------------------------
def _clasificar_label(txt):
    t = strip_acentos(txt or "").lower()
    if "objeto" in t:                                            return "objeto"
    if "adquirid" in t or "activos adquir" in t or "vendedor" in t or "transferid" in t:
        return "objeto"
    if "comprador" in t or "adquirent" in t or "adquiriente" in t:
        return "comprador"
    return None

_RE_GRUPO = re.compile(r"^(grupo|empresas?\s+del|lado|parte\s+(compradora|adquirente|vendedora))\b")
_RE_PAREN_ROL = re.compile(r"\(\s*(?:grupo\s+|empresa\s+)?"
                           r"(objeto|comprador(?:a)?|adquirente|adquiriente|vendedor)\s*\)", re.I)
_RE_LEGAL = re.compile(
    r"\b(S\.?\s?A\.?(\.?[UICF]\.?){0,4}|S\.?\s?R\.?\s?L|SRL|SA[UB]?|LTDA?|LIMITADA|LIMITED|"
    r"B\.?V|N\.?V|INC|LLC|GMBH|LTD|CO\.?\s?KG|\bKG\b|S\.?A\.?S|SARL|\bOY\b|APS|PTE|"
    r"CORPORATION|COMPANY|HOLDINGS?|SAICF|SACIF|SACIFIA|S\.?C\.?A|SDN\.?\s?BHD|SPA|S\.?L\.?U?)\b",
    re.I)
# stops por PREFIJO (ninguna empresa real empieza así)
_STOP_PREFIX = ("total", "otros", "subtotal", "combinado", "participacion", "ventas totales",
                "importaciones", "exportaciones", "posicion", "periodo", "descripcion", "cuota")
# stops EXACTOS (palabras que sí pueden ENCABEZAR un nombre real, p.ej.
# "EMPRESA DISTRIBUIDORA... S.A.", por eso solo se descartan si son toda la celda)
_STOP_EXACT = ("empresa", "empresas", "empresas involucradas", "var", "ihh", "hhi")

def _rol_parentesis(txt):
    m = _RE_PAREN_ROL.search(strip_acentos(txt or ""))
    return _clasificar_label(m.group(1)) if m else None

def _es_prosa(txt):
    t = (txt or "").strip()
    if len(t) < 15 or " " not in t:
        return False
    letras = sum(c.isalpha() for c in t)
    return letras >= max(10, len(t) * 0.45)

def _es_num(txt):
    t = (txt or "").strip()
    return bool(t) and re.fullmatch(r"[\d\.\,\%\s\-\$\(\)]+", t) is not None

def _es_fila_participacion(cells):
    joined = strip_acentos(" ".join(cells)).lower()
    if re.search(r"\b(19|20)\d{2}\b", joined):
        return True
    if re.search(r"participaci|\bihh\b|\bhhi\b|volumen|valor\s*\(|"
                 r"total\s+mercado|total\s+pais|cuota|market share|var\.?\s*ihh", joined):
        return True
    nov = [c for c in cells if c.strip()]
    if len(nov) >= 2 and sum(_es_num(c) for c in nov) >= len(nov) - 1:
        return True
    return False

def _es_header_empresas(cells):
    joined = strip_acentos(" ".join(cells)).lower()
    return "actividad" in joined and (
        "empresa" in joined or "afectad" in joined or "involucrad" in joined or "razon social" in joined)

def _limpiar_empresa(s):
    s = s.replace("|", " ")
    s = re.sub(r"\(\s*(?:grupo\s+|empresa\s+)?(?:objeto|comprador[a]?|adquirente|adquiriente|vendedor)\s*\)",
               "", s, flags=re.I)
    s = re.sub(r"^[\-•–·\s]+", "", s)
    s = re.sub(r"\b\d{1,3}\.\s+", " ", s)                    # números de lista "30. "
    s = re.sub(r"\s+\d{1,3}\.\s*$", "", s)
    s = re.sub(r"(?<=[a-záéíóúñü])\d{1,2}\b", "", s)          # superíndices de nota al pie
    s = re.sub(r"(?<=\.)\s*\d{1,2}$", "", s)                  # nota al pie tras "S.A." ("S.A.6", "S.A. 5")
    return re.sub(r"\s+", " ", s).strip(" “”\"'-–·")

def _parece_empresa(nombre):
    if not nombre or len(nombre) < 3 or _es_num(nombre):
        return False
    t = strip_acentos(nombre).lower().strip()
    if t in _STOP_EXACT or any(t == w or t.startswith(w + " ") for w in _STOP_PREFIX):
        return False
    if _clasificar_label(nombre) and len(nombre) <= 22:
        return False
    if _RE_LEGAL.search(nombre):
        return True
    tokens = [x for x in nombre.split() if x]
    if len(tokens) >= 2 and sum(1 for x in tokens if x[:1].isupper()) >= max(2, len(tokens) * 0.5):
        return True
    return bool(nombre.isupper() and len(nombre) >= 4)

def _label_de_seccion(cell, es_unica):
    # un marcador de rol ENTRE PARÉNTESIS -> "(objeto)" -> es una empresa con su
    # rol al lado, NO una etiqueta de sección.
    if _RE_PAREN_ROL.search(strip_acentos(cell)):
        return None
    r = _clasificar_label(cell)
    t = strip_acentos(cell).lower().strip()
    if r:
        if len(cell) <= 55 or _RE_GRUPO.match(t):
            return r
    elif es_unica and len(cell) <= 55 and _RE_GRUPO.match(t):
        return "comprador"          # "Grupo XXX" sin palabra clave -> lado comprador
    return None

def ex_empresas(ctx):
    res, activo, rol_actual = [], False, None
    for tabla in ctx["tablas"]:
        for row in tabla:
            if not row:
                continue
            cells = [(c or "").replace("\n", " ").strip() for c in row]
            nov = [(i, c) for i, c in enumerate(cells) if c]
            if not nov:
                continue
            if _es_header_empresas(cells):
                activo, rol_actual = True, None
                continue
            hay_prosa = any(_es_prosa(c) for _, c in nov)
            if not hay_prosa and _es_fila_participacion(cells):
                activo = False
                continue
            first_i, first_c = nov[0]
            lbl = _label_de_seccion(first_c, len(nov) == 1)
            if lbl:
                rol_actual, activo = lbl, True
                resto = nov[1:]
                if not resto:
                    continue                       # fila-etiqueta pura
                name_c = resto[0][1]
                act_cells = [c for _, c in resto[1:]]
            else:
                name_c = first_c
                act_cells = [c for _, c in nov[1:]]
            if not activo:
                continue
            nombre = _limpiar_empresa(name_c)
            valido = any(_es_prosa(c) for c in act_cells) or bool(_RE_LEGAL.search(nombre))
            if not valido or not _parece_empresa(nombre):
                continue
            rol = _rol_parentesis(name_c) or rol_actual
            if rol:
                res.append({"rol": rol, "nombre": nombre})

    vistos, out = set(), []
    for e in res:
        k = (e["rol"], strip_acentos(e["nombre"]).lower())
        if k not in vistos:
            vistos.add(k)
            out.append(e)
    return json.dumps(out, ensure_ascii=False) if out else None

EXTRACTORES = [
    ("tipo",                 ex_tipo),
    ("Carpeta",              ex_carpeta),
    ("Carátula",             ex_caratula),
    ("Grupo/Empresa",        ex_grupo),
    ("Empresas involucradas",ex_empresas),
    ("Fecha_Ingreso",        ex_fecha_ingreso),
    ("Fecha_firma",          ex_fecha_firma),
    ("Decisión",             ex_decision),
    ("Número de Resolución", ex_num_resolucion),
    ("Número de Dictamen",   ex_num_dictamen),
    ("Mercados relevantes",  ex_mercados),
    ("Relaciones económicas",ex_relaciones),
]

# ----------------------------------------------------------------------------
# Procesamiento de un PDF
# ----------------------------------------------------------------------------
def procesar_pdf(path):
    raw, tablas = leer_pdf(path)
    reso_raw, dict_raw = segmentar(raw)
    ctx = {
        "raw": raw, "flat": flat(raw),
        "reso_raw": reso_raw, "reso_flat": flat(reso_raw),
        "dict_raw": dict_raw, "dict_flat": flat(dict_raw),
        "tablas": tablas,
        "archivo": os.path.basename(path),
    }
    fila = {"Excluible": None}
    for nombre, fn in EXTRACTORES:
        try:
            val = fn(ctx)
        except Exception as e:
            val = None
            print(f"    [warn] {ctx['archivo']}: fallo en '{nombre}': {e}")
        fila[nombre] = val
        if nombre == "Carátula":       # la carátula alimenta al grupo/empresa
            ctx["caratula"] = val
    return fila

# ----------------------------------------------------------------------------
# Escritura de la planilla
# ----------------------------------------------------------------------------
def escribir_excel(filas, salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "firmadas"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri")
    for name, c in COL.items():
        cell = ws.cell(1, c, name)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for r, fila in enumerate(filas, start=2):
        for name in COLUMNS:
            if name == "meses":
                continue
            c = COL[name]
            val = fila.get(name)
            cell = ws.cell(r, c, val if val is not None else None)
            if name in ("Fecha_Ingreso", "Fecha_firma") and isinstance(val, datetime.date):
                cell.number_format = "DD/MM/YYYY"
            if name in ("Carátula", "Mercados relevantes", "Grupo/Empresa", "Empresas involucradas"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        # meses = fórmula (Excel la calcula al abrir); en blanco si falta alguna fecha
        fi, ff = get_column_letter(COL["Fecha_Ingreso"]), get_column_letter(COL["Fecha_firma"])
        ws.cell(r, COL["meses"],
                f'=IF(AND({fi}{r}<>"",{ff}{r}<>""),DATEDIF({fi}{r},{ff}{r},"m"),"")')

    anchos = {"tipo":8,"Carpeta":14,"Excluible":10,"Carátula":46,"Grupo/Empresa":30,
              "Empresas involucradas":60,
              "Fecha_Ingreso":14,"Fecha_firma":13,"Decisión":26,"Número de Resolución":26,
              "Número de Dictamen":30,"Mercados relevantes":50,"Relaciones económicas":20,"meses":8}
    for name, w in anchos.items():
        ws.column_dimensions[get_column_letter(COL[name])].width = w

    wb.save(salida)

# ----------------------------------------------------------------------------
def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else "pdfs"
    salida = sys.argv[2] if len(sys.argv) > 2 else "res_firmadas_nuevo.xlsx"
    pdfs = sorted(glob.glob(os.path.join(carpeta, "*.pdf")))
    if not pdfs:
        print(f"No se encontraron PDFs en '{carpeta}'.")
        sys.exit(1)

    filas, resumen = [], []
    for p in pdfs:
        print(f"[proc] {os.path.basename(p)}")
        try:
            fila = procesar_pdf(p)
        except Exception as e:
            print(f"    [ERROR] no se pudo procesar: {e}")
            continue
        filas.append(fila)
        vacios = [n for n in COLUMNS if n not in ("Excluible","meses") and not fila.get(n)]
        resumen.append((fila.get("Carpeta") or os.path.basename(p), vacios))

    # ordenar por número de CONC
    def keyn(f):
        m = re.search(r"(\d+)", f.get("Carpeta") or "")
        return int(m.group(1)) if m else 0
    filas.sort(key=keyn)

    escribir_excel(filas, salida)
    print(f"\nOK -> {salida}  ({len(filas)} filas)")
    print("\nCampos en blanco por caso (para revisar a mano):")
    for carpeta_id, vacios in resumen:
        print(f"  {carpeta_id}: {', '.join(vacios) if vacios else '— completo —'}")

if __name__ == "__main__":
    main()

