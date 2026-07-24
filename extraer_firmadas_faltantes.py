#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor de dictámenes CNDC SOLO para los PDFs que faltan analizar.

Reutiliza EXACTAMENTE la misma lógica de extracción de
`extraer_firmadas2.py` (se importa como módulo, no se copia código), pero en
lugar de procesar toda una carpeta, procesa únicamente los PDFs de
`<buscador>/pdf` que:

  1. NO figuran en el Excel `firm.xlsx` (columna 'Carpeta'), y
  2. NO pertenecen a la carpeta de 2017
     (C:\\Users\\Admin\\Documents\\ANC\\descargas_cndc\\2017), que fueron
     excluidos a propósito.

La comparación se hace por tipo (CONC/OPI) + número de expediente, normalizando
los distintos formatos de nombre ("CONC-1234.pdf", "conc 1234.pdf",
"CONC.1234 (PROSUM)", ...).

Uso:
    python extraer_firmadas_faltantes.py
        -> procesa los faltantes de ./pdf y escribe firmadas_faltantes.xlsx

    python extraer_firmadas_faltantes.py <salida.xlsx>
        -> idem, con nombre de salida a elección

    python extraer_firmadas_faltantes.py --listar
        -> solo imprime la lista de faltantes detectados, sin procesar

Dependencias:
    pip install pymupdf openpyxl   (ya requeridas por el script original)
"""
import os, re, sys, importlib.util

# --- Rutas fijas del entorno ------------------------------------------------
BASE_BUSCADOR = r"C:\Users\Admin\Documents\GitHub\buscadorANC"
DIR_PDF       = os.path.join(BASE_BUSCADOR, "pdf")
EXCEL_BASE    = os.path.join(BASE_BUSCADOR, "firm.xlsx")
DIR_2017      = r"C:\Users\Admin\Documents\ANC\descargas_cndc\2017"
SCRIPT_ORIG   = r"C:\Users\Admin\Documents\ANC\descargas_cndc\extraer_firmadas2.py"


# --- Importar la lógica del script original como módulo ---------------------
def cargar_extractor():
    """Carga extraer_firmadas2.py como módulo sin ejecutar su main()
    (está protegido por if __name__ == '__main__')."""
    spec = importlib.util.spec_from_file_location("extraer_firmadas2", SCRIPT_ORIG)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# --- Normalización de nombres para comparar fuentes -------------------------
def clave(nombre):
    """Devuelve (tipo, numero) a partir de un nombre de archivo o de una
    entrada de la columna 'Carpeta'. tipo en {'CONC','OPI',None}; numero es el
    primer entero sin ceros a la izquierda. Ignora sufijos '_2' (mismo caso)."""
    u = nombre.upper()
    tipo = "CONC" if "CONC" in u else ("OPI" if "OPI" in u else None)
    m = re.search(r"\d+", u)
    if not tipo or not m:
        return None
    return (tipo, m.group(0).lstrip("0") or "0")


def numeros_de_carpeta(dir_path):
    claves = set()
    for f in os.listdir(dir_path):
        if f.lower().endswith(".pdf"):
            k = clave(f)
            if k:
                claves.add(k)
    return claves


def numeros_de_excel(path_xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(path_xlsx, read_only=True, data_only=True)
    ws = wb["firmadas"] if "firmadas" in wb.sheetnames else wb.active
    claves = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        carp = row[1] if len(row) > 1 else None   # columna 'Carpeta'
        if carp:
            k = clave(str(carp))
            if k:
                claves.add(k)
    wb.close()
    return claves


def pdfs_faltantes():
    """Lista de rutas de PDF en DIR_PDF que faltan analizar (no en Excel,
    no en 2017). Agrupa por caso, incluyendo variantes '_2' del mismo número."""
    en_excel = numeros_de_excel(EXCEL_BASE)
    en_2017  = numeros_de_carpeta(DIR_2017)
    faltan = []
    for f in os.listdir(DIR_PDF):
        if not f.lower().endswith(".pdf"):
            continue
        k = clave(f)
        if k is None:
            continue                      # archivos sueltos sin CONC/OPI -> se ignoran
        if k in en_excel or k in en_2017:
            continue
        faltan.append(os.path.join(DIR_PDF, f))
    # ordenar por (tipo, numero) y luego nombre (para que '_2' vaya junto)
    def keyf(p):
        k = clave(os.path.basename(p))
        return (k[0], int(k[1]), os.path.basename(p))
    faltan.sort(key=keyf)
    return faltan


# ---------------------------------------------------------------------------
def main():
    args = [a for a in sys.argv[1:]]
    solo_listar = "--listar" in args
    args = [a for a in args if a != "--listar"]
    salida = args[0] if args else os.path.join(BASE_BUSCADOR, "firmadas_faltantes.xlsx")

    faltan = pdfs_faltantes()
    print(f"PDFs faltantes detectados (no en firm.xlsx y no en 2017): {len(faltan)}\n")
    for p in faltan:
        print(f"  {os.path.basename(p)}")

    if solo_listar:
        return
    if not faltan:
        print("\nNo hay faltantes para procesar.")
        return

    ext = cargar_extractor()
    print(f"\nProcesando {len(faltan)} PDFs con la lógica de extraer_firmadas2.py ...\n")

    filas, resumen = [], []
    for p in faltan:
        print(f"[proc] {os.path.basename(p)}")
        try:
            fila = ext.procesar_pdf(p)
        except Exception as e:
            print(f"    [ERROR] no se pudo procesar: {e}")
            continue
        filas.append(fila)
        vacios = [n for n in ext.COLUMNS
                  if n not in ("Excluible", "meses") and not fila.get(n)]
        resumen.append((fila.get("Carpeta") or os.path.basename(p), vacios))

    # ordenar por número de CONC/OPI (misma clave que el original)
    def keyn(f):
        m = re.search(r"(\d+)", f.get("Carpeta") or "")
        return int(m.group(1)) if m else 0
    filas.sort(key=keyn)

    ext.escribir_excel(filas, salida)
    print(f"\nOK -> {salida}  ({len(filas)} filas)")
    print("\nCampos en blanco por caso (para revisar a mano):")
    for carpeta_id, vacios in resumen:
        print(f"  {carpeta_id}: {', '.join(vacios) if vacios else '— completo —'}")


if __name__ == "__main__":
    main()
