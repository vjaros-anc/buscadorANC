#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Extractor de dictámenes CNDC — VERSIÓN OCR (para PDFs ESCANEADOS).

Es la variante de `extraer_firmadas.py` pensada para dictámenes que NO tienen
capa de texto (escaneos como imagen). Cada página se rasteriza y se pasa por
Tesseract (OCR) para recuperar el texto; después se aplican los mismos
extractores por campo, con expresiones regulares tolerantes a las dos eras de
formato:
    - Ley 25.156 / Art. 8 (notificación) / Art. 13 (decisión)  -> códigos #MP
    - Ley 27.442 / Art. 8 (notificación) / Art. 14 (decisión)  -> códigos #MEC

Diseño (igual filosofía que la v1): cada campo se extrae de forma independiente
y devuelve None si no encuentra el dato de forma fiable; nunca se inventa. Un PDF
que falla no corta el lote.

Como el OCR entrega texto PLANO (sin geometría de tablas), "Mercados relevantes"
y "Empresas involucradas" se intentan primero desde tablas (si la página tuviera
capa de texto) y, si no, desde la prosa; si no hay certeza, quedan en blanco.

Requisitos:
    pip install pymupdf openpyxl pytesseract pillow
    + Tesseract-OCR instalado (Windows: build de UB Mannheim) con idioma español.
      Descargar 'spa.traineddata'. Si tesseract.exe no está en el PATH, se busca
      en rutas típicas o se puede fijar con la variable de entorno TESSERACT_CMD.

Uso:
cd "C:\Users\Admin\Documents\ANC\descargas_cndc\2018\escaneados"
python extraer_firmadas_ocr.py .

"""
import sys, os, re, glob, unicodedata, datetime, json
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# OCR (Tesseract) — configuración perezosa
# ----------------------------------------------------------------------------
OCR_DPI = 300          # resolución de rasterizado antes de OCR
OCR_LANG = "spa"       # idioma; requiere spa.traineddata instalado
_OCR_READY = None      # None=sin probar, True/False una vez resuelto

def _localizar_tesseract():
    """Devuelve la ruta a tesseract.exe o None. Prioriza TESSERACT_CMD."""
    env = os.environ.get("TESSERACT_CMD")
    if env and os.path.isfile(env):
        return env
    candidatos = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
        os.path.expandvars(r"%LOCALAPPDATA%\Tesseract-OCR\tesseract.exe"),
    ]
    for c in candidatos:
        if c and os.path.isfile(c):
            return c
    return None  # se asumirá que está en el PATH

def _init_ocr():
    """Importa pytesseract y fija la ruta del binario. Cachea el resultado."""
    global _OCR_READY, pytesseract, Image
    if _OCR_READY is not None:
        return _OCR_READY
    try:
        import pytesseract as _pt
        from PIL import Image as _Img
    except Exception as e:
        print("    [OCR] Falta 'pytesseract' o 'pillow'. Instalá:  "
              "pip install pytesseract pillow")
        _OCR_READY = False
        return False
    globals()["pytesseract"] = _pt
    globals()["Image"] = _Img
    ruta = _localizar_tesseract()
    if ruta:
        _pt.pytesseract.tesseract_cmd = ruta
    try:
        _pt.get_tesseract_version()
    except Exception:
        print("    [OCR] No se encontró el ejecutable de Tesseract. Instalalo "
              "(UB Mannheim) o fijá la variable TESSERACT_CMD con la ruta a "
              "tesseract.exe.")
        _OCR_READY = False
        return False
    _OCR_READY = True
    return True

def _ocr_pagina(pg):
    """Rasteriza una página y devuelve el texto reconocido por OCR (o '')."""
    if not _init_ocr():
        return ""
    pix = pg.get_pixmap(dpi=OCR_DPI)
    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    try:
        return pytesseract.image_to_string(img, lang=OCR_LANG) or ""
    except Exception as e:
        # idioma no instalado u otro problema -> reintenta con 'eng' como último recurso
        try:
            return pytesseract.image_to_string(img) or ""
        except Exception:
            print(f"    [OCR] fallo de reconocimiento: {e}")
            return ""

# ----------------------------------------------------------------------------
# Configuración de columnas (idéntica a la v1)
# ----------------------------------------------------------------------------
COLUMNS = [
    "tipo", "Carpeta", "Excluible", "Carátula", "Grupo/Empresa",
    "Empresas involucradas",
    "Fecha_Ingreso", "Fecha_firma", "Decisión",
    "Número de Resolución", "Número de Dictamen",
    "Mercados relevantes", "Relaciones económicas", "meses",
]
COL = {name: i + 1 for i, name in enumerate(COLUMNS)}

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

# ----------------------------------------------------------------------------
# Lectura del documento:  texto por página si existe, si no -> OCR
# ----------------------------------------------------------------------------
def leer_pdf(path, umbral_ocr=20):
    """Devuelve (texto_crudo, tablas). Usa la capa de texto cuando está; para
    páginas escaneadas (sin texto) recurre a OCR. Las tablas por geometría solo
    aparecen en páginas con capa de texto (en escaneos puros quedan vacías)."""
    partes, tablas, n_ocr = [], [], 0
    with fitz.open(path) as doc:
        for pg in doc:
            txt = pg.get_text(sort=True) or ""
            if len(txt.strip()) < umbral_ocr:
                txt = _ocr_pagina(pg)
                n_ocr += 1
            partes.append(txt)
            try:
                for tb in pg.find_tables().tables:
                    tablas.append(tb.extract())
            except Exception:
                pass
    raw = "\n".join(partes).replace("\xa0", " ")
    return raw, tablas, n_ocr

def flat(txt):
    return re.sub(r"\s+", " ", txt).strip()

def segmentar(raw):
    idx = raw.find("Dictamen firma conjunta")
    if idx == -1:
        idx = raw.find("Dictamen firma")
    if idx > 0:
        return raw[:idx], raw[idx:]
    return raw, raw

def strip_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", s or "")
                   if unicodedata.category(c) != "Mn")

def _na(s):
    """Quita acentos preservando la posición de cada carácter (1:1), para
    poder anclar regex sobre texto OCR con acentos corruptos (ó->é, í->i) y
    seguir usando los índices sobre el texto original."""
    out = []
    for c in (s or ""):
        d = strip_acentos(c)
        out.append(d if len(d) == 1 else c)
    return "".join(out)

def parse_fecha_es(dia, mes_txt, anio):
    m = MESES_ES.get(strip_acentos(mes_txt).lower())
    if not m:
        return None
    try:
        return datetime.date(int(anio), m, int(dia))
    except ValueError:
        return None

def limpiar(s):
    return re.sub(r"\s+", " ", s).strip(" “”\"'") if s else s

RE_FECHA = r"(\d{1,2}) de ([A-Za-zÁÉÍÓÚáéíóúÜü]+) de (\d{4})"

# ----------------------------------------------------------------------------
# Extractores por campo  (str/date  ó  None) — regex tolerantes a ambas eras
# ----------------------------------------------------------------------------
def ex_tipo(ctx):
    return "CONC"

def ex_carpeta(ctx):
    m = re.search(r"CONC\.?\s*N?[.°º]?\s*(\d{2,4})", ctx["raw"])
    return f"CONC. {m.group(1)}" if m else None

def ex_caratula(ctx):
    # admite "caratulado: «...»" con o sin dos puntos; ancla sin acentos y
    # recupera el texto del original (mismos índices por _na).
    na, orig = _na(ctx["dict_flat"]), ctx["dict_flat"]
    m = re.search(r"caratulad[oa]s?\s*:?\s*[\"«“]([^\"»”]{5,300})[\"»”]", na, re.I)
    if not m:
        return None
    s = orig[m.start(1):m.end(1)]
    s = re.sub(r"^\s*CONC\.?\s*(?:N[°º\.]?\s*)?\d+\s*[-–]\s*", "", s)          # prefijo
    s = re.sub(r"\s*[\(\[]\s*CONC\.?\s*(?:N[°º\.]?\s*)?\d+\s*[\)\]]\s*$", "", s)  # sufijo "(CONC N° 1003)"
    return limpiar(s)

def ex_grupo(ctx):
    car = ctx.get("caratula")
    if car:
        m = re.search(r"^(.*?)\s*S/\s*(?:NOTIFIC|SOLICIT)", car, re.I)
        if m and m.group(1).strip():
            return limpiar(m.group(1))
    m = re.search(r"adquisici[oó]n\s+(?:de control\s+)?por parte de\s+(?:la firma\s+)?"
                  r"([A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúñ0-9&.\-\s]+?)\s+"
                  r"(?:del|de la|de las|de los)\b", ctx["dict_flat"])
    return limpiar(m.group(1)) if m else None

def ex_fecha_ingreso(ctx):
    # anclas sin acentos (texto OCR): recibio/notificacion/notificada, etc.
    d = _na(ctx["dict_flat"])
    for pat in (
        r"El d[ií]a\s+" + RE_FECHA + r"[^.]{0,70}?(?:recibi[oó]|notific)",
        r"(?:recibi[oó] la notificaci[oó]n|notificaron|recibi[oó] la solicitud)[^.]{0,90}?" + RE_FECHA,
        r"[Cc]on fecha\s+" + RE_FECHA + r"[^.]{0,90}?(?:recibi[oó]|notific)",
        r"notificad[ao]s?\s+(?:con fecha\s+|el\s+)?" + RE_FECHA,
    ):
        m = re.search(pat, d)
        if m:
            f = parse_fecha_es(m.group(1), m.group(2), m.group(3))
            if f:
                return f
    return None

def ex_fecha_firma(ctx):
    # La fecha de firma es la de "CIUDAD DE BUENOS AIRES + fecha" que acompaña a
    # la RESOLUCIÓN. Según el formato, la resolución puede ir primero (2023) o al
    # final (escaneos 2018). Si hay varias fechas, se elige la más cercana a un
    # marcador de resolución (RESUELVE / RESOL- / "Resolución").
    txt = _na(ctx["flat"])
    fechas = list(re.finditer(r"CIUDAD DE BUENOS AIRES\s+(?:[A-Za-z]+\s+)?" + RE_FECHA, txt))
    if not fechas:
        return None
    if len(fechas) > 1:
        marc = [mm.start() for mm in
                re.finditer(r"RESUELVE|RESOL-\d{4}|\bResolucion\b", txt)]
        if marc:
            fechas.sort(key=lambda f: min(abs(f.start() - p) for p in marc))
    m = fechas[0]
    return parse_fecha_es(m.group(1), m.group(2), m.group(3))

def ex_decision(ctx):
    """Detecta sentido + inciso del artículo de decisión (Art. 13 de la Ley
    25.156 ó Art. 14 de la Ley 27.442) y emite la forma canónica:
        "<Sentido> Art.<n> <letra>) Ley <ley>"
    """
    d = ctx["dict_flat"]
    sentido = letra = art = ley = None
    m = re.search(r"(Autoriza|Rechaza|Subordina|Deniega|Condiciona)\s*"
                  r"[^.\n]{0,30}?Art\.?\s*(1[34])[^.\n]*?\(?\s*([abc])\s*\)", d, re.I)
    if m:
        sentido, art, letra = m.group(1).capitalize(), m.group(2), m.group(3).lower()
    else:
        inc = re.search(r"art[íi]culo\s*(1[34])\s*[,\s]*inc(?:iso)?\.?\s*\(?([abc])\)?",
                        d, re.I)
        if inc:
            art, letra = inc.group(1), inc.group(2).lower()
            sentido = {"a": "Autoriza", "b": "Subordina", "c": "Rechaza"}[letra]
    if not (sentido and letra):
        return None
    lm = re.search(r"Ley\s*N?[°º.]?\s*(2[0-9]\.\d{3})", d)
    ley = lm.group(1) if lm else ("27.442" if art == "14" else "25.156")
    return f"{sentido} Art.{art} {letra}) Ley {ley}"

def ex_num_resolucion(ctx):
    # RESOL-2018-101-APN-SECC#MP  /  RESOL-2023-..-APN-SC#MEC  (org y sufijo variables)
    m = re.search(r"RESOL-\d{4}-\d+-APN-[A-Z0-9]+#\w+", ctx["raw"])
    return m.group(0) if m else None

def ex_num_dictamen(ctx):
    m = re.search(r"IF-\s*\d{4}-\s*\d+-APN-CNDC#\w+", ctx["dict_raw"]) \
        or re.search(r"IF-\s*\d{4}-\s*\d+-APN-CNDC#\w+", ctx["raw"])
    return re.sub(r"\s+", "", m.group(0)) if m else None

# ----------------------------------------------------------------------------
# Mercados relevantes (tabla de "Objeto" — solo si hay tablas con capa de texto)
# ----------------------------------------------------------------------------
def _celda_numerica(s):
    s = (s or "").strip()
    return bool(s) and re.fullmatch(r"[\d\.,%\s\-\$]+", s) is not None

def _label_objeto(s):
    # admite "Objeto", "Empresa Objeto" y "Empresas Objeto" (singular/plural)
    return bool(re.match(r"^\s*(empresas?\s+)?objeto\b", s.strip(), re.I)) and "(" not in s

def _mercados_desde_tablas(ctx):
    descs = []
    for tabla in ctx["tablas"]:
        recolectando = False
        for row in tabla:
            if not row:
                continue
            first = (row[0] or "").strip()
            rest = [(c or "").strip() for c in row[1:]]
            rest_txt = " ".join(x for x in rest if x).strip()
            if first and not rest_txt:
                recolectando = _label_objeto(first)
                continue
            if recolectando:
                desc = " ".join(x for x in rest if x and not _celda_numerica(x)).strip()
                if desc and re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]{3,}", desc):
                    descs.append(re.sub(r"\s+", " ", desc).strip())
    return descs

# núcleo del nombre de una empresa: primeras palabras antes de la forma jurídica
# (p.ej. "EMULGRAIN S.A." -> "EMULGRAIN"; "MOLINOS RIO DE LA PLATA S.A." -> "MOLINOS RIO DE LA PLATA")
def _nucleo_nombre(nombre):
    n = re.split(r"\s+(?:S\.?\s?A\.?|S\.?\s?R\.?\s?L|SA[UB]?|SRL|B\.?V|N\.?V|LLC|INC|LTD|"
                 r"S\.?A\.?C|LIMITADA|GMBH|SARL|SPA)\b", nombre, maxsplit=1, flags=re.I)[0]
    return n.strip()

def _mercados_desde_prosa(ctx):
    """En escaneos (tabla = imagen) toma la 'actividad económica principal' de la
    empresa OBJETO desde su descripción en prosa: 'X ... dedicada a ...' /
    'X ... tiene por objeto ...' / 'cuya actividad ... es ...'."""
    objetos = [e["nombre"] for e in _lista_empresas(ctx) if e["rol"] == "objeto"]
    if not objetos:
        return []
    na, orig = _na(ctx["dict_flat"]), ctx["dict_flat"]
    descs = []
    for nom in objetos:
        nucleo = re.escape(_na(_nucleo_nombre(nom)))
        # patrón definitorio: "X (es una sociedad/empresa ...) dedicada a / tiene por objeto ..."
        m = re.search(nucleo + r"\b(?:\s+S\.?A\.?[\w.]*)?\s+es una (?:sociedad|empresa)\b"
                      r".{0,250}?(?:dedicad[ao]s?\s+|tiene por objeto\s+|"
                      r"cuya actividad\w*\s+(?:es|consiste en)\s+)"
                      r"(.{5,300}?)(?:\.\s|\.$|;)", na, re.I | re.S)
        if not m:
            continue
        desc = orig[m.start(1):m.end(1)]
        desc = re.sub(r"^al?\s+(?:la\s+|el\s+|los\s+|las\s+)?", "", desc, flags=re.I)  # "al/a la ..."
        desc = re.sub(r"\s+", " ", desc).strip(" .,;")
        if re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]{3,}", desc):
            descs.append(desc)
    return descs

def ex_mercados(ctx):
    descs = _mercados_desde_tablas(ctx) or _mercados_desde_prosa(ctx)
    if not descs:
        return None
    return "; ".join(descs)[:400].strip(" ;")

def ex_relaciones(ctx):
    d = ctx["dict_flat"]
    has_h = re.search(r"relaci[oó]n(?:es)? horizontal(?:es)?|"
                      r"presenta relaciones horizontales|"
                      r"fortalecimiento (?:del grupo|horizontal)|"
                      r"relaci[oó]n horizontal", d, re.I) is not None
    has_v = re.search(r"efectos? vertical(?:es)?|relaci[oó]n(?:es)? vertical(?:es)?|"
                      r"reforzamiento (?:de\s+)?(?:un |las |los )?(?:efectos?|relaci[oó]n(?:es)?)?\s*vertical",
                      d, re.I) is not None
    has_c = re.search(r"conglomerado", d, re.I) is not None
    partes = [p for p, ok in (("h", has_h), ("v", has_v), ("c", has_c)) if ok]
    if not partes:
        return None
    nombre = {"h": "Horizontal", "v": "Vertical", "c": "Conglomerado"}
    if len(partes) == 1:
        return nombre[partes[0]]
    if set(partes) == {"h", "v"}:
        return "Horizontal + vertical"
    return " + ".join(nombre[p] for p in partes)

# ----------------------------------------------------------------------------
# Empresas involucradas
#   1) si hay tablas con capa de texto -> mismo método que la v1 (tabla).
#   2) si no (escaneo puro) -> intento por prosa de la descripción de la operación.
#   Salida: JSON [{"rol","nombre"}, ...] preservando el orden, o None.
# ----------------------------------------------------------------------------
def _clasificar_label(txt):
    t = strip_acentos(txt or "").lower()
    if "objeto" in t:                                            return "objeto"
    if "adquirid" in t or "activos adquir" in t or "vendedor" in t or "transferid" in t:
        return "objeto"
    if "comprador" in t or "adquirent" in t or "adquiriente" in t:
        return "comprador"
    return None

_RE_GRUPO = re.compile(r"^(grupo|empresas?\s+del|lado|parte\s+(compradora|adquirente|vendedora))\b")
_RE_PAREN_ROL = re.compile(r"\(\s*(?:grupo\s+|empresa\s+)?"
                           r"(objeto|comprador(?:a)?|adquirente|adquiriente|vendedor)\s*\)", re.I)
_RE_LEGAL = re.compile(
    r"\b(S\.?\s?A\.?(\.?[UICF]\.?){0,4}|S\.?\s?R\.?\s?L|SRL|SA[UB]?|LTDA?|LIMITADA|LIMITED|"
    r"B\.?V|N\.?V|INC|LLC|GMBH|LTD|CO\.?\s?KG|\bKG\b|S\.?A\.?S|SARL|\bOY\b|APS|PTE|"
    r"CORPORATION|COMPANY|HOLDINGS?|SAICF|SACIF|SACIFIA|S\.?A\.?C\.?I\.?F\.?I|S\.?C\.?A|"
    r"SDN\.?\s?BHD|SPA|S\.?L\.?U?)\b", re.I)
_STOP_PREFIX = ("total", "otros", "subtotal", "combinado", "participacion", "ventas totales",
                "importaciones", "exportaciones", "posicion", "periodo", "descripcion", "cuota")
_STOP_EXACT = ("empresa", "empresas", "empresas involucradas", "var", "ihh", "hhi")

def _rol_parentesis(txt):
    m = _RE_PAREN_ROL.search(strip_acentos(txt or ""))
    return _clasificar_label(m.group(1)) if m else None

def _es_prosa(txt):
    t = (txt or "").strip()
    if len(t) < 15 or " " not in t:
        return False
    letras = sum(c.isalpha() for c in t)
    return letras >= max(10, len(t) * 0.45)

def _es_num(txt):
    t = (txt or "").strip()
    return bool(t) and re.fullmatch(r"[\d\.\,\%\s\-\$\(\)]+", t) is not None

def _es_fila_participacion(cells):
    joined = strip_acentos(" ".join(cells)).lower()
    if re.search(r"\b(19|20)\d{2}\b", joined):
        return True
    if re.search(r"participaci|\bihh\b|\bhhi\b|volumen|valor\s*\(|"
                 r"total\s+mercado|total\s+pais|cuota|market share|var\.?\s*ihh", joined):
        return True
    nov = [c for c in cells if c.strip()]
    if len(nov) >= 2 and sum(_es_num(c) for c in nov) >= len(nov) - 1:
        return True
    return False

def _es_header_empresas(cells):
    joined = strip_acentos(" ".join(cells)).lower()
    return "actividad" in joined and (
        "empresa" in joined or "afectad" in joined or "involucrad" in joined or "razon social" in joined)

def _limpiar_empresa(s):
    s = s.replace("|", " ")
    s = re.sub(r"\(\s*(?:grupo\s+|empresa\s+)?(?:objeto|comprador[a]?|adquirente|adquiriente|vendedor)\s*\)",
               "", s, flags=re.I)
    s = re.sub(r"^[\-•–·\s]+", "", s)
    s = re.sub(r"\b\d{1,3}\.\s+", " ", s)
    s = re.sub(r"\s+\d{1,3}\.\s*$", "", s)
    s = re.sub(r"(?<=[a-záéíóúñü])\d{1,2}\b", "", s)
    s = re.sub(r"(?<=\.)\s*\d{1,2}$", "", s)                  # nota al pie tras "S.A." ("S.A.6", "S.A. 5")
    return re.sub(r"\s+", " ", s).strip(" “”\"'-–·")

def _parece_empresa(nombre):
    if not nombre or len(nombre) < 3 or _es_num(nombre):
        return False
    t = strip_acentos(nombre).lower().strip()
    if t in _STOP_EXACT or any(t == w or t.startswith(w + " ") for w in _STOP_PREFIX):
        return False
    if _clasificar_label(nombre) and len(nombre) <= 22:
        return False
    if _RE_LEGAL.search(nombre):
        return True
    tokens = [x for x in nombre.split() if x]
    if len(tokens) >= 2 and sum(1 for x in tokens if x[:1].isupper()) >= max(2, len(tokens) * 0.5):
        return True
    return bool(nombre.isupper() and len(nombre) >= 4)

def _label_de_seccion(cell, es_unica):
    # un marcador de rol ENTRE PARÉNTESIS -> "(objeto)" -> es una empresa con su
    # rol al lado, NO una etiqueta de sección.
    if _RE_PAREN_ROL.search(strip_acentos(cell)):
        return None
    r = _clasificar_label(cell)
    t = strip_acentos(cell).lower().strip()
    if r:
        if len(cell) <= 55 or _RE_GRUPO.match(t):
            return r
    elif es_unica and len(cell) <= 55 and _RE_GRUPO.match(t):
        return "comprador"
    return None

def _empresas_desde_tablas(tablas):
    res, activo, rol_actual = [], False, None
    for tabla in tablas:
        for row in tabla:
            if not row:
                continue
            cells = [(c or "").replace("\n", " ").strip() for c in row]
            nov = [(i, c) for i, c in enumerate(cells) if c]
            if not nov:
                continue
            if _es_header_empresas(cells):
                activo, rol_actual = True, None
                continue
            hay_prosa = any(_es_prosa(c) for _, c in nov)
            if not hay_prosa and _es_fila_participacion(cells):
                activo = False
                continue
            first_c = nov[0][1]
            lbl = _label_de_seccion(first_c, len(nov) == 1)
            if lbl:
                rol_actual, activo = lbl, True
                resto = nov[1:]
                if not resto:
                    continue
                name_c = resto[0][1]
                act_cells = [c for _, c in resto[1:]]
            else:
                name_c = first_c
                act_cells = [c for _, c in nov[1:]]
            if not activo:
                continue
            nombre = _limpiar_empresa(name_c)
            valido = any(_es_prosa(c) for c in act_cells) or bool(_RE_LEGAL.search(nombre))
            if not valido or not _parece_empresa(nombre):
                continue
            rol = _rol_parentesis(name_c) or rol_actual
            if rol:
                res.append({"rol": rol, "nombre": nombre})
    return res

# --- fallback por prosa (escaneos sin tabla) ---------------------------------
# Reconoce el patrón típico de la descripción de la operación:
#   "... adquisición ... de [acciones/capital de] OBJETO ... por (parte de) COMPRADOR(es)
#    ... [a las vendedoras VENDEDOR(es)] ..."
# el \s antes del sufijo evita matchear "SA" pegado dentro de palabras
# (p.ej. OLEAGINO-SA), que fragmentaba los nombres en el texto OCR.
_RE_EMP_TOK = (r"[A-ZÁÉÍÓÚÑ][\wÁÉÍÓÚÑáéíóúñ&\.\- ]*?\s"
               r"(?:S\.?\s?A\.?(?:\.?[UICF]\.?){0,4}|S\.?\s?R\.?\s?L\.?|S\.?A\.?C\.?I\.?F\.?I\.?"
               r"|S\.?L\.?|B\.?V\.?|LLC|INC\.?|LTD\.?|LIMITADA|GMBH|N\.?V\.?)")

def _nombres_en(fragmento):
    if not fragmento:
        return []
    out, vistos = [], set()
    for m in re.finditer(_RE_EMP_TOK, fragmento):
        nom = _limpiar_empresa(m.group(0))
        if _parece_empresa(nom):
            k = strip_acentos(nom).lower()
            if k not in vistos:
                vistos.add(k)
                out.append(nom)
    return out

def _empresas_desde_prosa(dict_flat):
    # Ancla sobre texto sin acentos (OCR) pero extrae los nombres del original.
    na = _na(dict_flat)
    m = re.search(r"consiste en\s+(.{20,700}?)"
                  r"(?:Como consecuencia|Con fecha|Tras el|A los efectos|\. [A-Z][a-z]|$)",
                  na, re.I)
    if not m:
        return []
    a, b = m.span(1)
    frase_na, frase_or = na[a:b], dict_flat[a:b]      # _na preserva índices

    # OBJETO: "(acciones|capital|participación|control|activos) ... de X (por|a favor de)"
    mo = re.search(r"(?:acciones|capital(?:\s+social)?|participaci[o]n|control|activos)\s+"
                   r"(?:social\s+)?(?:y\s+votos\s+)?de\s+(.{3,150}?)\s+(?:por\b|a favor de\b)",
                   frase_na, re.I)
    obj_or = frase_or[mo.start(1):mo.end(1)] if mo else None

    # COMPRADOR: "por (parte de) X" hasta el bloque de vendedores / a favor / fin
    mc = re.search(r"\bpor\s+(?:parte de\s+)?(.{3,220}?)"
                   r"(?:,?\s*a\s+(?:las?|los?)\s+vendedor|,?\s*de\s+(?:las?|los?)\s+vendedor|"
                   r",?\s*a favor|$)", frase_na, re.I)
    comp_or = frase_or[mc.start(1):mc.end(1)] if mc else None

    res, vistos = [], set()
    for rol, frag in (("objeto", obj_or), ("comprador", comp_or)):
        for n in _nombres_en(frag):
            k = (rol, strip_acentos(n).lower())
            if k not in vistos:
                vistos.add(k)
                res.append({"rol": rol, "nombre": n})
    return res

def _lista_empresas(ctx):
    if "_empresas" in ctx:
        return ctx["_empresas"]
    res = _empresas_desde_tablas(ctx["tablas"])
    if not res:
        res = _empresas_desde_prosa(ctx["dict_flat"])
    vistos, out = set(), []
    for e in res:
        k = (e["rol"], strip_acentos(e["nombre"]).lower())
        if k not in vistos:
            vistos.add(k)
            out.append(e)
    ctx["_empresas"] = out
    return out

def ex_empresas(ctx):
    out = _lista_empresas(ctx)
    return json.dumps(out, ensure_ascii=False) if out else None

EXTRACTORES = [
    ("tipo",                 ex_tipo),
    ("Carpeta",              ex_carpeta),
    ("Carátula",             ex_caratula),
    ("Grupo/Empresa",        ex_grupo),
    ("Empresas involucradas",ex_empresas),
    ("Fecha_Ingreso",        ex_fecha_ingreso),
    ("Fecha_firma",          ex_fecha_firma),
    ("Decisión",             ex_decision),
    ("Número de Resolución", ex_num_resolucion),
    ("Número de Dictamen",   ex_num_dictamen),
    ("Mercados relevantes",  ex_mercados),
    ("Relaciones económicas",ex_relaciones),
]

# ----------------------------------------------------------------------------
def procesar_pdf(path):
    raw, tablas, n_ocr = leer_pdf(path)
    reso_raw, dict_raw = segmentar(raw)
    ctx = {
        "raw": raw, "flat": flat(raw),
        "reso_raw": reso_raw, "reso_flat": flat(reso_raw),
        "dict_raw": dict_raw, "dict_flat": flat(dict_raw),
        "tablas": tablas,
        "archivo": os.path.basename(path),
    }
    fila = {"Excluible": None, "_ocr_pages": n_ocr}
    for nombre, fn in EXTRACTORES:
        try:
            val = fn(ctx)
        except Exception as e:
            val = None
            print(f"    [warn] {ctx['archivo']}: fallo en '{nombre}': {e}")
        fila[nombre] = val
        if nombre == "Carátula":
            ctx["caratula"] = val
    return fila

# ----------------------------------------------------------------------------
def escribir_excel(filas, salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "firmadas"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri")
    for name, c in COL.items():
        cell = ws.cell(1, c, name)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for r, fila in enumerate(filas, start=2):
        for name in COLUMNS:
            if name == "meses":
                continue
            c = COL[name]
            val = fila.get(name)
            cell = ws.cell(r, c, val if val is not None else None)
            if name in ("Fecha_Ingreso", "Fecha_firma") and isinstance(val, datetime.date):
                cell.number_format = "DD/MM/YYYY"
            if name in ("Carátula", "Mercados relevantes", "Grupo/Empresa", "Empresas involucradas"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        fi, ff = get_column_letter(COL["Fecha_Ingreso"]), get_column_letter(COL["Fecha_firma"])
        ws.cell(r, COL["meses"],
                f'=IF(AND({fi}{r}<>"",{ff}{r}<>""),DATEDIF({fi}{r},{ff}{r},"m"),"")')

    anchos = {"tipo":8,"Carpeta":14,"Excluible":10,"Carátula":46,"Grupo/Empresa":30,
              "Empresas involucradas":60,
              "Fecha_Ingreso":14,"Fecha_firma":13,"Decisión":26,"Número de Resolución":26,
              "Número de Dictamen":30,"Mercados relevantes":50,"Relaciones económicas":20,"meses":8}
    for name, w in anchos.items():
        ws.column_dimensions[get_column_letter(COL[name])].width = w

    # Si el archivo está abierto en Excel (bloqueado), guardar con nombre alternativo.
    try:
        wb.save(salida)
        return salida
    except PermissionError:
        base, ext = os.path.splitext(salida)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{stamp}{ext}"
        wb.save(alt)
        print(f"    [aviso] '{salida}' estaba abierto/bloqueado; se guardó como '{alt}'.")
        return alt

# ----------------------------------------------------------------------------
def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else "pdfs"
    salida = sys.argv[2] if len(sys.argv) > 2 else "res_firmadas_ocr.xlsx"
    pdfs = sorted(glob.glob(os.path.join(carpeta, "*.pdf")) +
                  glob.glob(os.path.join(carpeta, "*.PDF")))
    # de-duplicar (Windows es case-insensitive)
    vistos, unicos = set(), []
    for p in pdfs:
        k = os.path.normcase(os.path.abspath(p))
        if k not in vistos:
            vistos.add(k); unicos.append(p)
    pdfs = unicos
    if not pdfs:
        print(f"No se encontraron PDFs en '{carpeta}'.")
        sys.exit(1)

    filas, resumen = [], []
    for p in pdfs:
        print(f"[proc] {os.path.basename(p)}")
        try:
            fila = procesar_pdf(p)
        except Exception as e:
            print(f"    [ERROR] no se pudo procesar: {e}")
            continue
        if fila.get("_ocr_pages"):
            print(f"    [OCR] {fila['_ocr_pages']} página(s) reconocida(s) por OCR")
        filas.append(fila)
        vacios = [n for n in COLUMNS if n not in ("Excluible","meses") and not fila.get(n)]
        resumen.append((fila.get("Carpeta") or os.path.basename(p), vacios))

    def keyn(f):
        m = re.search(r"(\d+)", f.get("Carpeta") or "")
        return int(m.group(1)) if m else 0
    filas.sort(key=keyn)

    salida = escribir_excel(filas, salida)
    print(f"\nOK -> {salida}  ({len(filas)} filas)")
    print("\nCampos en blanco por caso (para revisar a mano):")
    for carpeta_id, vacios in resumen:
        print(f"  {carpeta_id}: {', '.join(vacios) if vacios else '— completo —'}")

if __name__ == "__main__":
    main()
